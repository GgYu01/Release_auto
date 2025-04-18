import os
from datetime import datetime
from logging import Logger
from typing import Dict, Optional, List, Any
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from utils.git_utils import GitOperator
from config.schemas import AllReposConfig, ExcelConfig, GitRepoInfo, CommitDetail


class ExcelReporter:
    def __init__(self, logger: Logger, git_operator: GitOperator, excel_config: ExcelConfig):
        self.logger: Logger = logger
        self.git_operator: GitOperator = git_operator
        self.excel_config: ExcelConfig = excel_config

    def generate_report(self, all_repos_config: AllReposConfig, version_info: Dict[str, Any], target_excel_path: str) -> bool:
        if not self.excel_config.enabled:
            self.logger.info("Excel report generation is disabled in the configuration.")
            return True

        self.logger.info(f"Starting Excel report generation: {target_excel_path}")

        workbook = None
        worksheet = None
        existing_data: List[List[Any]] = []
        created_new_workbook: bool = False

        try:
            if os.path.exists(target_excel_path):
                self.logger.info(f"Loading existing workbook: {target_excel_path}")
                try:
                    workbook = openpyxl.load_workbook(target_excel_path)
                    worksheet = workbook.active
                    for row_index in range(2, worksheet.max_row + 1):
                        row_data = [cell.value for cell in worksheet[row_index]]
                        existing_data.append(row_data)
                    self.logger.info(f"Read {len(existing_data)} existing data rows from the workbook.")
                except InvalidFileException:
                    self.logger.warning(f"File {target_excel_path} exists but is not a valid Excel file. Creating a new workbook.")
                    workbook = openpyxl.Workbook()
                    worksheet = workbook.active
                    created_new_workbook = True
                except Exception as e_load:
                    self.logger.error(f"Error loading existing workbook {target_excel_path}: {e_load}. Creating a new workbook.", exc_info=True)
                    workbook = openpyxl.Workbook()
                    worksheet = workbook.active
                    created_new_workbook = True
            else:
                self.logger.info(f"Creating a new workbook: {target_excel_path}")
                workbook = openpyxl.Workbook()
                worksheet = workbook.active
                created_new_workbook = True

            if not worksheet:
                self.logger.error("Failed to obtain an active worksheet.")
                return False

            if created_new_workbook:
                headers = ["Release Version", "Commit Message", "Module", "Patch Path", "Column E",
                           "Zircon/Garnet Commits", "Tester/Author/MTK", "Date", "Col I", "Col J",
                           "Col K", "Col L", "Col M", "Col N", "Commit ID"]
                worksheet.append(headers)
                self.logger.info("Added headers to the new workbook.")

            release_version: str = version_info.get('latest_tag', 'N/A')
            current_date: str = datetime.now().strftime('%Y/%m/%d')

            zircon_repo: Optional[GitRepoInfo] = next((repo for repo in all_repos_config.all_git_repos() if repo.repo_name == self.excel_config.zircon_repo_name), None)
            garnet_repo: Optional[GitRepoInfo] = next((repo for repo in all_repos_config.all_git_repos() if repo.repo_name == self.excel_config.garnet_repo_name), None)

            zircon_commit_id: str = "N/A"
            if zircon_repo:
                repo_path_z = getattr(zircon_repo, 'repo_path', None)
                local_branch_z = getattr(zircon_repo, 'local_branch', None)
                if repo_path_z:
                    zircon_commit_id = self.git_operator.get_latest_commit_id(repo_path_z, local_branch_z) or "N/A"
                    if zircon_commit_id == "N/A":
                         self.logger.warning(f"Could not retrieve latest commit ID for Zircon repo: {zircon_repo.repo_name}")
                else:
                    self.logger.warning(f"Zircon repo '{zircon_repo.repo_name}' found but missing 'repo_path'.")
            else:
                 self.logger.warning(f"Zircon repo '{self.excel_config.zircon_repo_name}' not found in configuration.")

            garnet_commit_id: str = "N/A"
            if garnet_repo:
                 repo_path_g = getattr(garnet_repo, 'repo_path', None)
                 local_branch_g = getattr(garnet_repo, 'local_branch', None)
                 if repo_path_g:
                     garnet_commit_id = self.git_operator.get_latest_commit_id(repo_path_g, local_branch_g) or "N/A"
                     if garnet_commit_id == "N/A":
                         self.logger.warning(f"Could not retrieve latest commit ID for Garnet repo: {garnet_repo.repo_name}")
                 else:
                      self.logger.warning(f"Garnet repo '{garnet_repo.repo_name}' found but missing 'repo_path'.")
            else:
                 self.logger.warning(f"Garnet repo '{self.excel_config.garnet_repo_name}' not found in configuration.")

            commit_string_f: str = f"zircon:{zircon_commit_id}\ngarnet:{garnet_commit_id}"

            new_rows_data: List[List[Any]] = []
            for repo_info in all_repos_config.all_git_repos():
                if not repo_info.commit_details:
                    continue
                for commit in repo_info.commit_details:
                    commit_module_val: str = ""
                    try:
                        if commit.commit_module and len(commit.commit_module) > 0:
                            commit_module_val = commit.commit_module[0]
                        elif repo_info.repo_parent:
                             commit_module_val = repo_info.repo_parent
                    except IndexError:
                         self.logger.warning(f"IndexError accessing commit_module for commit {commit.id}, using repo_parent '{repo_info.repo_parent}'.")
                         commit_module_val = repo_info.repo_parent if repo_info.repo_parent else ""
                    except Exception as e_mod:
                         self.logger.error(f"Unexpected error processing commit module for {commit.id}: {e_mod}", exc_info=True)
                         commit_module_val = repo_info.repo_parent if repo_info.repo_parent else ""

                    row_data = [
                        release_version,
                        commit.message,
                        commit_module_val,
                        commit.patch_path or "",
                        "",
                        commit_string_f,
                        f"{self.excel_config.tester_name} / {commit.author} / {self.excel_config.mtk_owner_serial}",
                        current_date,
                        "是",
                        "是，76未测试",
                        "",
                        "",
                        "",
                        "",
                        commit.id
                    ]
                    new_rows_data.append(row_data)

            if new_rows_data:
                self.logger.info(f"Preparing to insert {len(new_rows_data)} new rows into the Excel sheet.")
                try:
                    worksheet.insert_rows(idx=2, amount=len(new_rows_data))
                    self.logger.info(f"Successfully inserted {len(new_rows_data)} blank rows at index 2.")

                    for row_index, row_content in enumerate(new_rows_data, start=2):
                        for col_index, cell_value in enumerate(row_content, start=1):
                            worksheet.cell(row=row_index, column=col_index, value=cell_value)
                    self.logger.info(f"Successfully wrote {len(new_rows_data)} new data rows starting at row 2.")

                except Exception as e_insert:
                    self.logger.error(f"Error inserting or writing new rows: {e_insert}", exc_info=True)
                    return False

            else:
                self.logger.info("No new commit data found to add to the report.")

            if existing_data:
                start_row_for_old_data = 2 + len(new_rows_data)
                self.logger.info(f"Preparing to write back {len(existing_data)} existing rows starting at row {start_row_for_old_data}.")
                try:
                    for row_index, row_content in enumerate(existing_data, start=start_row_for_old_data):
                        for col_index, cell_value in enumerate(row_content, start=1):
                             if col_index <= len(row_content):
                                 worksheet.cell(row=row_index, column=col_index, value=cell_value)
                    self.logger.info(f"Successfully wrote back {len(existing_data)} existing rows.")
                except Exception as e_write_old:
                    self.logger.error(f"Error writing back existing data: {e_write_old}", exc_info=True)


            self.logger.info(f"Saving workbook to {target_excel_path}")
            try:
                workbook.save(target_excel_path)
                self.logger.info("Excel report generated successfully.")
                return True
            except PermissionError:
                self.logger.error(f"Permission denied when trying to save {target_excel_path}.", exc_info=True)
                return False
            except Exception as e_save: # Catch other potential save errors
                 self.logger.error(f"An error occurred while saving the workbook {target_excel_path}: {e_save}", exc_info=True)
                 return False

        except Exception as e_main:
            self.logger.error(f"An unexpected error occurred during Excel report generation: {e_main}", exc_info=True)
            return False